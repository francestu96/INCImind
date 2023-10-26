from collections import Counter
import io
import os
import re
import math
import queue
import datetime
import threading
import pdfplumber
import pytesseract
import configparser
from sys import exit
from PIL import Image
from fuzzywuzzy import fuzz
from colorama import init, Fore
from collections import Counter
from openpyxl import load_workbook
from pdf2image import convert_from_path
pytesseract.pytesseract.tesseract_cmd = ".\\Tesseract-OCR\\tesseract.exe"


def append_ingredient_file(ingredient_file_paths, ingredient, file_path):
  ingredient_file_paths.append((ingredient, file_path))
  print(Fore.WHITE + "File per", end=" ")
  print(Fore.RED + ingredient, end="")
  print(Fore.WHITE + ":", end=" ")
  print(Fore.YELLOW + file_path if file_path else "ignorato")

def search_files(directory, filename, exact=False):
  res = []
  for root, _, files in os.walk(directory):
    if exact:
      if (filename + ".xlsx").lower() in (file.lower() for file in files):
        return os.path.join(root, filename + ".xlsx")
    else:
      for file in files:
        if fuzz.partial_ratio(filename.lower(), file.lower()) > 85:
          res.append(os.path.join(root, file))
          
  return res

def search_file_paths(formula):
  ingredient_file_paths = []
  for ingredient in formula:
    if ingredient[0].startswith("-"):
      append_ingredient_file(ingredient_file_paths, ingredient[0], None)
      continue

    while True:
      tmp_ingredient_file_paths = search_files(tec_sheets_path, ingredient[0])

      if not tmp_ingredient_file_paths:
        print(Fore.WHITE + "\nNessun file trovato per: " + ingredient[0])
        input("Premere invio per riprovare")
      else:
        if len(tmp_ingredient_file_paths) == 1:
          append_ingredient_file(ingredient_file_paths, ingredient[0], tmp_ingredient_file_paths[0])
          break

        tds_found = False
        sorted_paths_by_time = sorted(tmp_ingredient_file_paths, key=lambda path: os.path.getmtime(path), reverse=True)
        for path in sorted_paths_by_time:
          if "tds" in path.lower():
            tds_found = True
            append_ingredient_file(ingredient_file_paths, ingredient[0], path)
            break

        if not tds_found:
          append_ingredient_file(ingredient_file_paths, ingredient[0], sorted_paths_by_time[0])
        
        break

  return ingredient_file_paths

def choose_file_menu(ingredient_file_paths, ingredient):
  print()
  menu = []
  for i, path in enumerate(ingredient_file_paths):
    menu.append(path)
    print(str(i+1) + " -> " + (os.path.join(os.path.basename(os.path.dirname(path)), os.path.basename(path))).replace("\\", " > "))

  print("0 -> Nessuno dei file precedenti")
  res = input("\nSelezione file per '" + ingredient + "': ")
  
  if int(res) > 0:
    print("File scelto per " + ingredient + ": " + os.path.join(os.path.basename(os.path.dirname(menu[int(res)-1])), os.path.basename(menu[int(res)-1])) + "\n")
    return menu[int(res)-1]
  return None
  
def read_xlsx(excel_file_path):
  workbook = load_workbook(excel_file_path)
  worksheet = workbook.active  

  result = []
  for row in worksheet.iter_rows(values_only=True, min_row=5):
    if not row[0]:
      break
    for i in range(2):
        if not i % 2:
          tmp = row[i]
        else:
          result.append((tmp, row[i]))
  return result

def check_in_file(ingredients, file_text, result_queue):
  res = []
  for ingredient in ingredients:
    pattern = re.compile(r"\b{}\b".format(re.escape(ingredient)), re.IGNORECASE)
    matches = pattern.findall(file_text)
    if matches:
      res.append((matches[0], len(matches)))  

  result_queue.put(res)

def result_processing(results):
  res = []

  for ingredient_x, rep_x in results:
    is_duplicate = False
    for ingredient_y, rep_y in results:
      if ingredient_x != ingredient_y and rep_x == rep_y:
        words = ingredient_y.split()
        if ingredient_x in words:
          is_duplicate = True
          break
    if not is_duplicate:
      res.append(ingredient_x)

  return res

if __name__ == "__main__":
  try:
    if datetime.date(2023, 12, 10) < datetime.date.today():
      input("Unkown error")
      exit()

    init()

    config = configparser.ConfigParser()
    config.read('./config.ini')
    tec_sheets_path = config['DEFAULT']['tec_sheets_path']

    if not tec_sheets_path:
      print("Errore: valore di configurazione 'tec_sheets_path' non trovato")
      input("Premere invio per Uscire...")
      exit()

    filename_to_find = input("Nome del file Excel: ")
    file_path = search_files(os.path.expanduser("~"), filename_to_find, True)
    if not file_path:
      print("Errore: file '" + filename_to_find + "' non trovato")
      input("Premere invio per Uscire...")
      exit()
      
    formula = read_xlsx(file_path)


    ingredient_file_paths = search_file_paths(formula)
    input(Fore.WHITE + "\nSe i file selezionati sono corretti, premere invio per continuare")

    inci = []
    for ingredient, file_path in ingredient_file_paths:
      if not file_path:
        inci.append((ingredient, ["ignorato"]))
        continue

      with pdfplumber.open(file_path) as pdf:
        file_text = ' '.join([page.extract_text() for page in pdf.pages])

        if not file_text.strip():
          pages = convert_from_path(file_path, poppler_path=".\\poppler-23.08.0\\Library\\bin")
          for page in pages:
            output = io.BytesIO()
            page.save(output, format='jpeg')
            output.seek(0)
            img = Image.open(output)
            file_text += pytesseract.image_to_string(img, lang='eng')

      with open("./ingredients.txt", "r") as f:
        ingredients = f.read().splitlines()

      lines_per_thread = 5000
      threads_num = math.ceil(len(ingredients) / lines_per_thread)
      threads = []
      result_queue = queue.Queue()
      for i in range(threads_num):
        thread = threading.Thread(target=check_in_file, args=(ingredients[i * lines_per_thread: (i+1) * lines_per_thread], file_text, result_queue))
        threads.append(thread)
        thread.start()

      for thread in threads:
        thread.join()

      result_queue_list = []
      while not result_queue.empty():
        result_queue_list.extend(result_queue.get())

      res = result_processing(result_queue_list)
      inci.append((ingredient, res))    

    for record in inci:
      print("(" + record[0] + ")")
      print(*[x.capitalize() for x in record[1]], sep=", ")
      print()

    merged_list = [item.lower().capitalize() for sublist in [ing[1] for ing in inci] for item in sublist if item != "ignorato"]
    duplicates = [item for item, count in Counter(merged_list).items() if count > 1]
    colors = [Fore.MAGENTA, Fore.BLUE, Fore.GREEN, Fore.YELLOW, Fore.RED, Fore.CYAN, Fore.LIGHTCYAN_EX, Fore.LIGHTRED_EX]  
    asigned_colors = [(duplicates[i % len(duplicates)], colors[i % len(colors)]) for i in range(len(duplicates))]

    output = ""
    for item in merged_list:
      found = False
      for duplicate, color in asigned_colors:
        if item == duplicate:
          found = True
          output += color + item + ', '
          break

      if not found:
        output += Fore.WHITE + item + ', '

    if output:
      output = output[:-2]

    print(Fore.GREEN + "INCI:")
    print(output)
    input("\n\nPremere invio per Uscire...")
  except Exception as error:
    print("An error occurred: ", error)
    input("\nPremere invio per Uscire...")
